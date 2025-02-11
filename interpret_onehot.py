import os
import pickle
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd

from pipeline._utils import CHARPROTSET, cluster_mapping
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from skfeature.function.similarity_based import lap_score
from skfeature.utility import construct_W
from sklearn.feature_selection import VarianceThreshold

from matplotlib.ticker import FuncFormatter




def plot_positional_weights(mean_pw, std_pw, capped_indices, parent_dir):

    pos = range(len(mean_pw))
    avg_mw = mean_pw
    std_mw = std_pw
    # color = [weight["color"] for weight in weights]



    fig, axes = plt.subplots(6, 1, figsize=(14,10))

    start, stop = 0, 100
    for ax in axes:
        for x, y, err in zip(pos[start:stop], avg_mw[start:stop], std_mw[start:stop]):
            if capped_indices is None or x in capped_indices:
                ax.errorbar(x, y, yerr=err, color= "saddlebrown", fmt="o", capsize=5)  # data points
            ax.axvline(x, color='grey', linestyle='--', linewidth=0.5, alpha=0.6)  # Add vertical dashed lines

        # ax.set_ylim(0, 0.4)
        ax.set_ylim(0, 0.4)
        ax.set_xticks(np.arange(start, stop, 10))
        ax.axhline(0, color='black', linestyle='--', linewidth=0.8)  # Add a horizontal line at y=0
        ax.margins(x=0.01)
        # set the y-ticklabels to be the y_limit values
        ax.set_yticks(np.arange(0, 0.5, 0.2))


        # Set tick label fontsize
        ax.tick_params(axis='both', which='major', labelsize=18)
        ax.xaxis.set_major_formatter(FuncFormatter(lambda x, _: int(x) + 1))

        start += 100
        stop += 100

    # axes[0].set_title("Model Weights with Error Bars", fontsize=18)
    axes[5].set_xlabel("Position", fontsize=24)
    axes[3].set_ylabel("Weight", fontsize=24)

    plt.tight_layout()
    plt.savefig(os.path.join(parent_dir, "positional_weights.pdf"), dpi=300, format="pdf", bbox_inches="tight")
    plt.show()


# ================= plot positional influence figure =================
# parent_dir = "main-exp-result/trial-result-VLP200-ridge-onehotprotein-seqlen1426-optimfold9"
# # threshold = 0.12
# positional_weights, mean_pw, std_pw = get_positional_weights(parent_dir)
# # capped_indices = get_capped_indices(mean_pw, std_pw, threshold)
# plot_positional_weights(mean_pw, std_pw, capped_indices=None, parent_dir=parent_dir)
# ===================================================================


def get_aa_weights(parent_dir):
    with open(os.path.join(parent_dir, "misc.pkl"), "rb") as file:
        misc = pickle.load(file)
        model_weights = misc["model_weights"]

    total_weights = []
    # unpack the weights
    for model_number, weight in enumerate(model_weights):
        reshaped_weight = weight.reshape(1426, -1)  # shape: (max_length, #AAs)
        total_weights.append(reshaped_weight)   # shape: (#models, max_length, #AAs)

    total_weights = np.stack(total_weights, axis=0)   # shape: (#models, max_length, #AAs)

    mean_aw = total_weights.mean(axis=0)   # shape: (max_length, #AAs)
    std_aw = total_weights.std(axis=0)   # shape: (max_length, #AAs)
    return mean_aw, std_aw




# def plot_aa_weights(mean_aw, std_aw, parent_dir):
#     # Amino acid labels
#     CHARPROTSET_AA = np.array([k for k, v in CHARPROTSET.items()])
#
#     # Create a workbook and sheet
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "AA Weights"
#
#     # Header row
#     num_positions = mean_aw.shape[0]
#     for col in range(num_positions):
#         ws.cell(row=1, column=col + 1, value=f"Position {col + 1}")
#
#     # Number of rows
#     num_rows = 25
#
#     # Populate the data
#     for pos in range(num_positions):
#         weights = mean_aw[pos]
#
#         # Sort weights and standard deviations
#         sorted_indices = np.argsort(weights)[::-1]
#         sorted_weights = weights[sorted_indices]
#         sorted_stds = std_aw[pos][sorted_indices]
#         sorted_labels = CHARPROTSET_AA[sorted_indices]
#
#         for row in range(num_rows):
#             if row < len(sorted_labels):
#                 # Format cell content
#                 aa_label = sorted_labels[row]
#                 weight = sorted_weights[row]
#                 std = sorted_stds[row]
#                 cell_content = f"{aa_label}: {weight:.4f} ± {std:.4f}"
#
#                 # Determine color
#                 if weight > 0:
#                     fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Light pink
#                 elif weight < 0:
#                     fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue
#                 else:
#                     fill = None
#             else:
#                 cell_content = ""
#                 fill = None
#
#             # Write cell content
#             cell = ws.cell(row=row + 2, column=pos + 1, value=cell_content)
#
#             # Apply color
#             if fill:
#                 cell.fill = fill
#
#     # Save the workbook
#     output_path = f"{parent_dir}/aa_weights_colored.xlsx"
#     wb.save(output_path)
#     print(f"Excel file with colors saved at: {output_path}")


def plot_aa_weights(mean_aw, std_aw, parent_dir):
    # Amino acid labels
    CHARPROTSET_AA = np.array([k for k, v in CHARPROTSET.items()])

    # Create a workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "AA Weights"

    # Header row with amino acid labels
    num_positions = mean_aw.shape[0]
    num_rows = 25  # Number of amino acids to include per position

    ws.cell(row=1, column=1, value="Position")
    for col in range(num_rows):
        ws.cell(row=1, column=col + 2, value=f"AA {col + 1}")

    # Populate the data
    for pos in range(num_positions):
        weights = mean_aw[pos]

        # Sort weights and standard deviations
        sorted_indices = np.argsort(weights)[::-1]
        sorted_weights = weights[sorted_indices]
        sorted_stds = std_aw[pos][sorted_indices]
        sorted_labels = CHARPROTSET_AA[sorted_indices]

        # Write position as the first column of the row
        ws.cell(row=pos + 2, column=1, value=f"Position {pos + 1}")

        for col in range(num_rows):
            if col < len(sorted_labels):
                # Format cell content
                aa_label = sorted_labels[col]
                weight = sorted_weights[col]
                std = sorted_stds[col]
                cell_content = f"{aa_label}: {weight:.4f} ± {std:.4f}"

                # Determine color
                if weight > 0:
                    fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Light pink
                elif weight < 0:
                    fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue
                else:
                    fill = None
            else:
                cell_content = ""
                fill = None

            # Write cell content
            cell = ws.cell(row=pos + 2, column=col + 2, value=cell_content)

            # Apply color
            if fill:
                cell.fill = fill

    # Save the workbook
    wb.save(f"{parent_dir}/aa_weights_colored.xlsx")
    return

# ================= produce aa weights excel =================
# parent_dir = "main-exp-result/trial-result-VLP200-ridge-onehotprotein-seqlen1426-optimfold9"
# mean_aw, std_aw = get_aa_weights(parent_dir)
# plot_aa_weights(mean_aw, std_aw, parent_dir)
# ============================================================

def heatmap_aa_weights(mean_aw, std_aw, parent_dir):
    # Amino acid labels
    CHARPROTSET_AA = np.array([k for k, v in CHARPROTSET.items()])

    num_positions = 10  # Number of positions to include  mean_aw.shape[0]
    num_cols = 25  # Number of amino acids to include per position

    heatmap_data = np.full((num_positions, num_cols), np.nan)  # Initialize with NaNs
    annotations = np.full((num_positions, num_cols), '', dtype=object)  # For annotations

    aa_labels = []  # Store labels for x-axis

    for pos in range(num_positions):
        weights = mean_aw[pos]

        # Sort weights and standard deviations
        sorted_indices = np.argsort(weights)[::-1]
        sorted_weights = weights[sorted_indices]
        sorted_stds = std_aw[pos][sorted_indices]
        sorted_labels = CHARPROTSET_AA[sorted_indices]

        for col in range(min(num_cols, len(sorted_labels))):
            heatmap_data[pos, col] = sorted_weights[col]
            annotations[pos, col] = sorted_labels[col]  # Store label as annotation

        if pos == 0:
            aa_labels = [sorted_labels[col] for col in range(min(num_cols, len(sorted_labels)))]

    plt.figure(figsize=(7, 8))
    ax = sns.heatmap(heatmap_data, cmap='coolwarm', annot=annotations, fmt='', xticklabels=False,
                     yticklabels=[f"Position {i + 1}" for i in range(num_positions)],
                     annot_kws={"fontsize": 12}, cbar_kws={'orientation': 'horizontal'})
    ax.set_yticklabels(ax.get_yticklabels(), fontsize=12)  # Change yticklabel font size

    cbar = ax.collections[0].colorbar
    cbar.ax.tick_params(labelsize=12)  # Change color scale's font size
    # cbar.set_ticks(np.linspace(-0.03, 0.06, num=5))  # Add min and max to the color scale

    plt.xlabel("Amino Acids", fontsize=14)
    # plt.title("Amino Acid Weights Heatmap")
    plt.tight_layout()
    plt.savefig(f"{parent_dir}/aa_weights_heatmap.pdf", format="pdf", dpi=300, bbox_inches="tight")
    plt.show()


# ================= produce aa weights heatmap =================
# parent_dir = "main-exp-result/trial-result-VLP200-ridge-onehotprotein-seqlen1426-optimfold9"
# mean_aw, std_aw = get_aa_weights(parent_dir)
# heatmap_aa_weights(mean_aw, std_aw, parent_dir)
# ============================================================


def plot_cluster_weights(mean_aw, std_aw, parent_dir):
    # Amino acid labels
    cluster_keys = np.array(list(set([v for k, v in cluster_mapping.items()])))
    # Create a workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "AA Weights"

    # Header row with amino acid labels
    num_positions = mean_aw.shape[0]
    num_rows = 25  # Number of amino acids to include per position

    ws.cell(row=1, column=1, value="Position")
    for col in range(num_rows):
        ws.cell(row=1, column=col + 2, value=f"AA {col + 1}")

    # Populate the data
    for pos in range(num_positions):
        weights = mean_aw[pos]

        # Sort weights and standard deviations
        sorted_indices = np.argsort(weights)[::-1]
        sorted_weights = weights[sorted_indices]
        sorted_stds = std_aw[pos][sorted_indices]
        sorted_labels = cluster_keys[sorted_indices]

        # Write position as the first column of the row
        ws.cell(row=pos + 2, column=1, value=f"Position {pos + 1}")

        for col in range(num_rows):
            if col < len(sorted_labels):
                # Format cell content
                aa_label = sorted_labels[col]
                if aa_label == 1:
                    aa_label = "Aliphatic"
                elif aa_label == 2:
                    aa_label = "Aromatic"
                elif aa_label == 3:
                    aa_label = "Neutral"
                elif aa_label == 4:
                    aa_label = "Positively charged"
                elif aa_label == 5:
                    aa_label = "Negatively charged"
                elif aa_label == 6:
                    aa_label = "Special Cases"
                weight = sorted_weights[col]
                std = sorted_stds[col]
                cell_content = f"{aa_label}: {weight:.4f} ± {std:.4f}"

                # Determine color
                if weight > 0:
                    fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Light pink
                elif weight < 0:
                    fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue
                else:
                    fill = None
            else:
                cell_content = ""
                fill = None

            # Write cell content
            cell = ws.cell(row=pos + 2, column=col + 2, value=cell_content)

            # Apply color
            if fill:
                cell.fill = fill

    # Save the workbook
    wb.save(f"{parent_dir}/cluster_weights_colored.xlsx")
    return

# ================= produce cluster weights =================
# parent_dir = "study3-cluster/trial-result-VLP200-onehot-cluster-ridge-seqlen1426-optimfold9"
# mean_aw, std_aw=get_aa_weights(parent_dir)
# plot_cluster_weights(mean_aw, std_aw, parent_dir)
# ============================================================


def heatmap_cluster_weights(mean_aw, std_aw, parent_dir):
    # Cluster labels
    cluster_keys = np.array(list(set([v for k, v in cluster_mapping.items()])))

    num_positions = 10
    num_rows = 6  # Number of amino acids to include per position

    heatmap_data = np.full((num_positions, num_rows), np.nan)  # Initialize with NaNs
    annotations = np.full((num_positions, num_rows), '', dtype=object)  # For annotations

    for pos in range(num_positions):
        weights = mean_aw[pos]

        # Sort weights and standard deviations
        sorted_indices = np.argsort(weights)[::-1]
        sorted_weights = weights[sorted_indices]
        sorted_labels = cluster_keys[sorted_indices]

        for col in range(min(num_rows, len(sorted_labels))):
            heatmap_data[pos, col] = sorted_weights[col]

            # Assign human-readable cluster labels
            if sorted_labels[col] == 1:
                annotations[pos, col] = "Aliphatic"
            elif sorted_labels[col] == 2:
                annotations[pos, col] = "Aromatic"
            elif sorted_labels[col] == 3:
                annotations[pos, col] = "Neutral"
            elif sorted_labels[col] == 4:
                annotations[pos, col] = "Positively\ncharged"
            elif sorted_labels[col] == 5:
                annotations[pos, col] = "Negatively\ncharged"
            elif sorted_labels[col] == 6:
                annotations[pos, col] = "Special\nCases"

    plt.figure(figsize=(7, 8))  # Increase figure size to accommodate longer annotations
    ax = sns.heatmap(heatmap_data, cmap='coolwarm', annot=annotations,
                     fmt='', annot_kws={"size": 12},
                     xticklabels=False, yticklabels=[f"Position {i+1}" for i in range(num_positions)], cbar_kws={'orientation': 'horizontal'})



    ax.set_yticklabels(ax.get_yticklabels(), fontsize=12)  # Change yticklabel font size

    cbar = ax.collections[0].colorbar
    cbar.ax.tick_params(labelsize=12)  # Change color scale's font size
    # cbar.set_ticks([np.nanmin(heatmap_data), np.nanmax(heatmap_data)])  # Add min and max to the color scale

    plt.xlabel("Amino Acid Clusters", fontsize=14)
    plt.tight_layout()
    plt.savefig(f"{parent_dir}/cluster_weights_heatmap.pdf", format="pdf", dpi=300, bbox_inches="tight")
    plt.show()

# ================= produce cluster weights =================
# parent_dir = "study3-cluster/trial-result-VLP200-onehot-cluster-ridge-seqlen1426-optimfold9"
# mean_aw, std_aw=get_aa_weights(parent_dir)
# heatmap_cluster_weights(mean_aw, std_aw, parent_dir)
# ============================================================